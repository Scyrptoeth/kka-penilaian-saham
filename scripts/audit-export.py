#!/usr/bin/env python3
"""
Audit script for kka-template.xlsx — produces a punch list of gaps for
Session 024 (export coverage work).

Usage:
    python3 scripts/audit-export.py

Outputs an audit report to stdout (markdown). Pipe to file if needed.
"""
import json
import sys
import warnings
from pathlib import Path

from openpyxl import load_workbook

warnings.filterwarnings("ignore")

ROOT = Path(__file__).parent.parent
TEMPLATE = ROOT / "public/templates/kka-template.xlsx"

# Website nav-tree → expected Excel sheet name
# (some renaming because Excel uses different names)
WEBSITE_NAV = {
    "Input Master": [
        ("HOME", "HOME"),
    ],
    "Input Data": [
        ("Fixed Asset", "FIXED ASSET"),
        ("Balance Sheet", "BALANCE SHEET"),
        ("Income Statement", "INCOME STATEMENT"),
        ("Key Drivers", "KEY DRIVERS"),
        ("Acc Payables", "ACC PAYABLES"),
    ],
    "Analisis": [
        ("Financial Ratio", "FINANCIAL RATIO"),
        ("FCF", "FCF"),
        ("NOPLAT", "NOPLAT"),
        ("Growth Revenue", "GROWTH REVENUE"),
        ("ROIC", "ROIC"),
        ("Growth Rate", "GROWTH RATE"),
        ("Cash Flow Statement", "CASH FLOW STATEMENT"),
    ],
    "Proyeksi": [
        ("Proy. L/R", "PROY LR"),
        ("Proy. Fixed Asset", "PROY FIXED ASSETS"),
        ("Proy. Balance Sheet", "PROY BALANCE SHEET"),
        ("Proy. NOPLAT", "PROY NOPLAT"),
        ("Proy. Cash Flow", "PROY CASH FLOW STATEMENT"),
    ],
    "Penilaian": [
        ("DLOM", "DLOM"),
        ("DLOC (PFC)", "DLOC(PFC)"),
        ("WACC", "WACC"),
        ("Discount Rate", "DISCOUNT RATE"),
        ("Borrowing Cap", "BORROWING CAP"),
        ("DCF", "DCF"),
        ("AAM", "AAM"),
        ("EEM", "EEM"),
        ("CFI", "CFI"),
        ("Simulasi Potensi", "SIMULASI POTENSI (AAM)"),
    ],
    "Ringkasan": [
        ("Dashboard", "DASHBOARD"),
    ],
}

# Sheets in template that aren't in the website nav (possible cleanup candidates)
KNOWN_NON_NAV_SHEETS = {
    "DAFTAR EMITEN 2023": "DJP dataset (read-only reference)",
    "INCOME STATEMENT (2)": "Duplicate / legacy",
    "ADJUSTMENT TANAH": "Land adjustment helper (subset of FA)",
    "KUISIONER": "Legacy questionnaire",
    "PROY ACC PAYABLES": "Hidden helper for PROY CFS",
    "PROY FIXED ASSET": "Duplicate of PROY FIXED ASSETS (typo)",
    "PASAR PEMBANDING_": "Market comparable dataset",
    "PASAR KESEBANDINGAN1": "Market comparable dataset",
    "PASAR KESEBANDINGAN": "Market comparable dataset",
    "PASAR PEMBANDING": "Market comparable dataset",
    "PERBANDINGAN": "Market comparison helper",
    "HARGA PASAR SAHAM": "Stock price reference",
    "PANGSA PASAR": "Market share dataset",
    "TL": "Time/value helper sheet",
    "DIVIDEND DISCOUNT MODEL": "Alternate valuation (not in website)",
    "RESUME": "One-page summary (not in website nav)",
}


def count_cells(ws):
    formulas = 0
    values = 0
    for row in ws.iter_rows():
        for c in row:
            if c.value is None:
                continue
            if isinstance(c.value, str) and c.value.startswith("="):
                formulas += 1
            else:
                values += 1
    return formulas, values


def main():
    print("# Audit Report — Excel Export Coverage")
    print(f"\n**Template**: `{TEMPLATE.relative_to(ROOT)}`")
    print(f"\n**Generated**: by `scripts/audit-export.py`\n")

    wb = load_workbook(TEMPLATE, read_only=False)
    sheets_in_template = {}
    for name in wb.sheetnames:
        ws = wb[name]
        f, v = count_cells(ws)
        sheets_in_template[name] = {
            "state": ws.sheet_state,
            "formulas": f,
            "values": v,
        }

    # ─── Section 1: Website nav coverage ───
    print("## 1. Website Nav → Excel Sheet Coverage\n")
    print("| Website Nav | Expected Sheet | Found? | State | Formulas | Values | Verdict |")
    print("|---|---|---|---|---|---|---|")

    visibility_actions = []  # (sheet_name, current_state, target_state, reason)

    for group, items in WEBSITE_NAV.items():
        for label, sheet in items:
            info = sheets_in_template.get(sheet)
            if info is None:
                print(f"| {group} → {label} | `{sheet}` | ❌ NOT FOUND | — | — | — | **MISSING** |")
                continue
            verdict = "✓"
            if info["state"] == "hidden":
                verdict = "⚠️ Hidden — should unhide"
                visibility_actions.append((sheet, "hidden", "visible", f"in website nav ({group} → {label})"))
            print(
                f"| {group} → {label} | `{sheet}` | ✓ | {info['state']} | "
                f"{info['formulas']} | {info['values']} | {verdict} |"
            )

    # ─── Section 2: Non-nav sheets in template ───
    print("\n## 2. Non-Nav Sheets in Template (potential cleanup)\n")
    print("| Sheet | Current State | Formulas | Values | Annotation | Suggested |")
    print("|---|---|---|---|---|---|")

    nav_sheet_names = {sheet for items in WEBSITE_NAV.values() for _, sheet in items}
    for name, info in sheets_in_template.items():
        if name in nav_sheet_names:
            continue
        annotation = KNOWN_NON_NAV_SHEETS.get(name, "(unannotated)")
        if info["state"] == "visible":
            suggested = "**HIDE** (cleanup)"
            visibility_actions.append((name, "visible", "hidden", f"not in website nav: {annotation}"))
        else:
            suggested = "✓ already hidden"
        print(
            f"| `{name}` | {info['state']} | {info['formulas']} | {info['values']} | "
            f"{annotation} | {suggested} |"
        )

    # ─── Section 3: Visibility action summary ───
    print("\n## 3. Visibility Action Summary\n")
    if not visibility_actions:
        print("_No visibility changes needed._")
    else:
        unhide = [a for a in visibility_actions if a[2] == "visible"]
        hide = [a for a in visibility_actions if a[2] == "hidden"]
        print(f"### Sheets to UNHIDE ({len(unhide)})")
        for sheet, _, _, reason in unhide:
            print(f"- `{sheet}` — {reason}")
        print(f"\n### Sheets to HIDE ({len(hide)})")
        for sheet, _, _, reason in hide:
            print(f"- `{sheet}` — {reason}")

    # ─── Section 4: Dynamic catalog overflow analysis ───
    print("\n## 4. Dynamic Catalog Overflow Analysis\n")

    # Read catalog sizes from source (best effort)
    bs_catalog = ROOT / "src/data/catalogs/balance-sheet-catalog.ts"
    is_catalog = ROOT / "src/data/catalogs/income-statement-catalog.ts"
    fa_catalog = ROOT / "src/data/catalogs/fixed-asset-catalog.ts"

    def count_catalog_entries(path):
        if not path.exists():
            return None
        text = path.read_text()
        # Crude count of `id:` occurrences
        return text.count("id:")

    bs_count = count_catalog_entries(bs_catalog)
    is_count = count_catalog_entries(is_catalog)
    fa_count = count_catalog_entries(fa_catalog)

    print("| Catalog | Total Accounts | Template Capacity | Notes |")
    print("|---|---|---|---|")
    print(f"| Balance Sheet | ~{bs_count} | rows ~8-49 | Original 84 baseline; user adds via dropdown |")
    print(f"| Income Statement | ~{is_count} | rows ~6-39 | Original 41 baseline |")
    print(f"| Fixed Asset | ~{fa_count} | rows ~7-26 (×7 sub-blocks) | Original 20 baseline |")

    print(
        "\n**Current handling**: extended accounts go into RINCIAN NERACA detail "
        "sheet only (BS only, since session 018). Subtotal/total formulas in main "
        "sheets do NOT reference RINCIAN values — extended accounts invisible to "
        "downstream Excel calculations."
    )
    print(
        "\n**Required (per user decision)**: insert rows + auto-shift formulas in "
        "main sheets so all extended accounts flow into subtotals/totals natively."
    )

    # ─── Section 5: Inspection of mapping coverage ───
    print("\n## 5. Cell-Mapping Coverage by Sheet\n")

    cm_path = ROOT / "src/lib/export/cell-mapping.ts"
    cm_text = cm_path.read_text() if cm_path.exists() else ""

    print("| Sheet | Has Scalar Mapping? | Has Grid Mapping? | Has Array? | Has Dynamic Rows? |")
    print("|---|---|---|---|---|")
    for name in sorted(sheets_in_template.keys()):
        if name not in nav_sheet_names and sheets_in_template[name]["state"] == "hidden":
            continue
        has_scalar = f"'{name}'" in cm_text and "kind: 'scalar'" in cm_text
        # Crude — check if any mapping references this sheet
        has_any = f"excelSheet: '{name}'" in cm_text
        # Couldn't easily distinguish kinds without parsing — just mark presence
        mark = "✓" if has_any else "—"
        print(f"| `{name}` | {mark} | {mark} | {mark} | {mark} |")

    print(
        "\n_(Coverage column granularity: presence-only — for detailed kind breakdown, "
        "inspect `cell-mapping.ts` directly.)_"
    )

    # ─── Section 6: Final punch list ───
    print("\n## 6. Punch List — Concrete Actions for Session 024\n")
    print("### Visibility changes (Phase A — quick, low-risk)")
    print("Modify `export-xlsx.ts` to set `ws.state = 'visible'` / `'hidden'` per sheet:")
    for sheet, _, target, reason in visibility_actions:
        print(f"- `{sheet}` → **{target}** ({reason})")

    print("\n### Dynamic catalog row insertion (Phase B — complex)")
    print(
        "- Implement insert-row logic for BS / IS / FA when user accounts exceed "
        "template baseline."
    )
    print(
        "- Auto-shift all dependent formulas (subtotals, totals, cross-sheet refs)."
    )
    print(
        "- Drop RINCIAN NERACA (no longer needed — accounts flow natively into main sheet)."
    )
    print("- Apply same pattern to IS and FA dynamic catalogs.")

    print("\n### Acc Payables + Key Drivers full mapping (Phase C — verify)")
    print(
        "- These sheets are now visible per Phase A. Verify cell mappings cover all "
        "user-editable inputs from the website pages."
    )

    print("\n### Verification (Phase D)")
    print(
        "- Generate sample export with extended accounts. Open in Excel manually. "
        "Verify subtotals/totals correctly reflect ALL accounts, including extras."
    )
    print("- Confirm no broken formulas after row-insert.")
    print("- Confirm only website-nav sheets are visible.")


if __name__ == "__main__":
    main()
