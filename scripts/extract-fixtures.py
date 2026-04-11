#!/usr/bin/env python3
"""
Extract KKA Penilaian Saham xlsx into JSON fixtures for TDD.

Output:
  __tests__/fixtures/_index.json    -> sheet metadata (name, slug, visible, dimensions)
  __tests__/fixtures/<slug>.json    -> per-sheet dump of cells with value + formula + format

Each cell entry:
  {
    "addr": "B4",
    "row": 4,
    "col": 2,
    "value": 1234.56,         # computed value (read-only load)
    "formula": "=SUM(B1:B3)", # if cell has a formula (data-only load)
    "number_format": "#,##0",
    "data_type": "n"          # n=number, s=string, d=date, b=bool, f=formula
  }

We do TWO passes over the workbook:
 1. data_only=True  -> to get computed values (Excel's last cached result)
 2. data_only=False -> to get original formulas

Only cells that are non-empty are serialized. Merged cells are noted per sheet.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "kka-penilaian-saham.xlsx"
FIXTURES_DIR = ROOT / "__tests__" / "fixtures"

# Sheets that the spec marks as hidden-but-required (referenced by visible sheets).
HIDDEN_REQUIRED = {
    "ACC PAYABLES",
    "KEY DRIVERS",
    "PROY ACC PAYABLES",
    "ADJUSTMENT TANAH",
}


def slugify(name: str) -> str:
    """BALANCE SHEET -> balance-sheet, DLOC(PFC) -> dloc-pfc, PROY LR -> proy-lr."""
    s = name.lower().strip()
    s = re.sub(r"[\s\(\)\/\\]+", "-", s)
    s = re.sub(r"[^a-z0-9\-]", "", s)
    s = re.sub(r"-+", "-", s).strip("-")
    return s


def cell_dict(cell_value, cell_formula) -> dict:
    """Build a compact cell entry from two parallel cell references."""
    out: dict = {
        "addr": cell_value.coordinate,
        "row": cell_value.row,
        "col": cell_value.column,
        "value": _jsonable(cell_value.value),
        "data_type": cell_value.data_type,
    }
    # Formula is on the formula-pass workbook (data_only=False).
    if isinstance(cell_formula.value, str) and cell_formula.value.startswith("="):
        out["formula"] = cell_formula.value
    if cell_value.number_format and cell_value.number_format != "General":
        out["number_format"] = cell_value.number_format
    return out


def _jsonable(v):
    """Convert openpyxl values into JSON-serializable primitives."""
    if v is None:
        return None
    if isinstance(v, (int, float, str, bool)):
        return v
    # Dates / datetimes -> ISO string
    try:
        return v.isoformat()
    except AttributeError:
        return str(v)


def extract_sheet(ws_value, ws_formula) -> dict:
    """Serialize one worksheet — only non-empty cells."""
    cells = []
    for row_value, row_formula in zip(
        ws_value.iter_rows(), ws_formula.iter_rows()
    ):
        for cv, cf in zip(row_value, row_formula):
            if cv.value is None and cf.value is None:
                continue
            cells.append(cell_dict(cv, cf))

    merged = [str(r) for r in ws_value.merged_cells.ranges]

    return {
        "name": ws_value.title,
        "slug": slugify(ws_value.title),
        "visible": ws_value.sheet_state == "visible",
        "dimensions": {
            "max_row": ws_value.max_row,
            "max_col": ws_value.max_column,
            "max_col_letter": get_column_letter(ws_value.max_column) if ws_value.max_column else "A",
        },
        "merged_ranges": merged,
        "cells": cells,
    }


def main() -> int:
    if not XLSX_PATH.exists():
        print(f"ERROR: not found: {XLSX_PATH}", file=sys.stderr)
        return 1

    print(f"Loading: {XLSX_PATH.name}")
    # Pass 1: data_only=True -> cached computed values
    wb_values = load_workbook(XLSX_PATH, data_only=True, read_only=False)
    # Pass 2: data_only=False -> original formulas
    wb_formulas = load_workbook(XLSX_PATH, data_only=False, read_only=False)

    FIXTURES_DIR.mkdir(parents=True, exist_ok=True)

    index = []
    for name in wb_values.sheetnames:
        ws_v = wb_values[name]
        ws_f = wb_formulas[name]

        is_visible = ws_v.sheet_state == "visible"
        is_hidden_required = name in HIDDEN_REQUIRED

        if not is_visible and not is_hidden_required:
            # Skip hidden sheets we explicitly said to ignore.
            print(f"  skip (hidden, not required): {name}")
            continue

        print(f"  extracting: {name}")
        data = extract_sheet(ws_v, ws_f)
        data["required_hidden"] = (not is_visible) and is_hidden_required

        slug = data["slug"]
        out_path = FIXTURES_DIR / f"{slug}.json"
        out_path.write_text(
            json.dumps(data, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

        index.append(
            {
                "name": name,
                "slug": slug,
                "visible": is_visible,
                "required_hidden": data["required_hidden"],
                "dimensions": data["dimensions"],
                "cell_count": len(data["cells"]),
                "file": f"{slug}.json",
            }
        )

    index_path = FIXTURES_DIR / "_index.json"
    index_path.write_text(
        json.dumps(
            {
                "source": XLSX_PATH.name,
                "sheet_count": len(index),
                "sheets": index,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    print(f"\nExtracted {len(index)} sheets -> {FIXTURES_DIR}")
    print(f"Index: {index_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
